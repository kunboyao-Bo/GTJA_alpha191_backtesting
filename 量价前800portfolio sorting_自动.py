import pandas as pd
import numpy as np
from scipy import stats
import tushare as ts


# ─────────────────────────────────────────────
# 0. 股票池：半年更新一次的沪深300+中证500快照（800只）
# ─────────────────────────────────────────────

def load_pivot(path):
    df = pd.read_csv(path).set_index("trade_date")
    df.index = pd.to_datetime(df.index)  # 关键：把字符串index转成真正的datetime
    df = df.sort_index()  # 顺便确保按真实时间顺序排好
    return df


def get_semiannual_snapshots(token, start_date, end_date):
    """
    每半年（6月30日、12月31日附近）获取沪深300+中证500成分股并集，构成800只股票池

    参数
    ----
    token : str
        Tushare Pro token
    start_date, end_date : str
        覆盖的时间区间，如 '2022-01-01', '2025-12-31'

    返回
    ----
    dict: {snapshot_date(pd.Timestamp): set(ts_code)}
    """
    pro = ts.pro_api(token)
    start = pd.Timestamp(start_date)
    end = pd.Timestamp(end_date)

    anchors = []
    for y in range(start.year, end.year + 1):
        for md in ['06-30', '12-31']:
            d = pd.Timestamp(f"{y}-{md}")
            if start <= d <= end:
                anchors.append(d)

    snapshots = {}
    for anchor in anchors:
        window_start = (anchor - pd.Timedelta(days=30)).strftime('%Y%m%d')
        window_end = anchor.strftime('%Y%m%d')

        codes = set()
        for idx_code in ['000300.SH', '000905.SH']:
            df = pro.index_weight(index_code=idx_code, start_date=window_start, end_date=window_end)
            if df.empty:
                continue
            latest = df['trade_date'].max()
            codes |= set(df.loc[df['trade_date'] == latest, 'con_code'])

        if codes:
            snapshots[anchor] = codes
            print(f"{anchor.date()} 股票池（沪深300+中证500）: {len(codes)} 只")
        else:
            print(f"警告: {anchor.date()} 未取到成分股数据")

    return snapshots


def get_universe_for_date(date, snapshots):
    """取 date 当天或之前最近一个半年快照对应的股票池；若无更早快照，兜底取最早的未来快照"""
    date = pd.Timestamp(date)
    past_anchors = [a for a in snapshots if a <= date]

    if not past_anchors:
        all_anchors = sorted(snapshots.keys())
        return list(snapshots[all_anchors[0]]) if all_anchors else []

    latest_anchor = max(past_anchors)
    return list(snapshots[latest_anchor])


# ─────────────────────────────────────────────
# 1. 因子截面：从已算好的日频因子矩阵取一天截面
# ─────────────────────────────────────────────
def factor_group_from_panel(signal_date, factor_panel_df, universe_list):
    universe_list = [c for c in universe_list if c in factor_panel_df.columns]

    factor = factor_panel_df.loc[signal_date, universe_list].dropna()

    if len(factor) < 50:
        return pd.DataFrame(columns=['因子值', '分组'])

    # Alpha21只有-1和1，直接映射为组1和组5，跳过qcut
    unique_vals = factor.unique()
    if set(unique_vals).issubset({-1.0, 1.0}):
        group_map = {-1.0: 1, 1.0: 5}
        groups = factor.map(group_map)
        return pd.DataFrame({'因子值': factor, '分组': groups})

    ranks = factor.rank(method='first', pct=True)
    groups = pd.qcut(ranks, q=5, labels=[1, 2, 3, 4, 5])
    return pd.DataFrame({'因子值': factor, '分组': groups})


# ─────────────────────────────────────────────
# 2. 收益计算：支持日/周/月任意频率
# ─────────────────────────────────────────────
def calc_adj_returns(close_df, stocks, start, end):
    start = pd.Timestamp(start)
    end = pd.Timestamp(end)

    stocks = [s for s in stocks if s in close_df.columns]

    trade_dates = pd.DatetimeIndex(close_df.index)

    start_td = trade_dates[trade_dates >= start][0]
    end_td = trade_dates[trade_dates <= end][-1]

    if start_td >= end_td:
        print(f"警告：start {start_td.date()} >= end {end_td.date()}，跳过")
        return pd.DataFrame(columns=stocks)

    p_start = close_df.loc[start_td, stocks]
    p_end = close_df.loc[end_td, stocks]

    ret = p_end / p_start - 1

    return pd.DataFrame([ret.values], index=[end_td], columns=stocks)


# ─────────────────────────────────────────────
# 3. 分组收益汇总
# ─────────────────────────────────────────────
def calc_quantile_returns(group_df, ret_df):
    result = {}
    for q in [1, 2, 3, 4, 5]:
        stocks = group_df[group_df['分组'] == q].index
        stocks = stocks[stocks.isin(ret_df.columns)]
        result[q] = ret_df[stocks].mean(axis=1)
    return pd.DataFrame(result)


# ─────────────────────────────────────────────
# 4. 换仓日程生成：按频率自动生成 (signal_date, ret_start, ret_end)
# ─────────────────────────────────────────────
def build_schedule(trading_dates, start, end, freq='W'):
    td = pd.DatetimeIndex(sorted(trading_dates))
    td = td[(td >= start) & (td <= end)]

    if freq == 'D':
        rebalance_dates = td

    elif freq == 'W':
        td_series = pd.Series(td, index=td)
        rebalance_dates = pd.DatetimeIndex(
            td_series.groupby(td_series.index.to_period('W')).last().values
        )

    elif freq == 'ME':
        td_series = pd.Series(td, index=td)
        rebalance_dates = pd.DatetimeIndex(
            td_series.groupby(td_series.index.to_period('M')).last().values
        )

    else:
        raise ValueError(f"不支持的频率: {freq}")

    schedule = []
    for i, sig_date in enumerate(rebalance_dates[:-1]):
        next_td = td[td > sig_date]
        if len(next_td) == 0:
            continue
        ret_start = next_td[0]
        ret_end = rebalance_dates[i + 1]
        if ret_start > ret_end:
            continue
        schedule.append((sig_date, ret_start, ret_end))

    return schedule


# ─────────────────────────────────────────────
# 5. IC Decay 分析：先跑这个，确认换仓频率
# ─────────────────────────────────────────────
def calc_ic_decay(factor_panel_df, close_df, max_lag=20):
    common_stocks = factor_panel_df.columns.intersection(close_df.columns)
    common_dates = factor_panel_df.index.intersection(close_df.index)

    factor = factor_panel_df.loc[common_dates, common_stocks]
    close = close_df.loc[common_dates, common_stocks]

    records = []
    for lag in range(1, max_lag + 1):
        # 未来 lag 天的累计持有期收益（而非"未来第 lag 天"当天的单日收益）
        fwd_ret = close.shift(-lag) / close - 1

        ic_series = []
        for date in factor.index:
            f_row = factor.loc[date].dropna()
            r_row = fwd_ret.loc[date].reindex(f_row.index).dropna()
            common = f_row.index.intersection(r_row.index)
            if len(common) < 30:
                continue
            ic, _ = stats.spearmanr(f_row[common], r_row[common])
            ic_series.append(ic)

        ic_arr = np.array(ic_series)
        ic_mean = ic_arr.mean()
        ic_std = ic_arr.std()
        t_stat = ic_mean / (ic_std / np.sqrt(len(ic_arr))) if ic_std > 0 else np.nan
        p_val = 2 * (1 - stats.t.cdf(abs(t_stat), df=len(ic_arr) - 1))

        records.append({
            'lag': lag,
            'IC均值': round(ic_mean, 4),
            'IC_std': round(ic_std, 4),
            't统计量': round(t_stat, 3),
            'p值': round(p_val, 4),
            '显著': '✓' if p_val < 0.05 else ''
        })

    result = pd.DataFrame(records).set_index('lag')
    print("\n── IC Decay 分析 ──")
    print(result.to_string())

    sig_lags = result[result['显著'] == '✓'].index.tolist()
    if not sig_lags:
        print("\n建议：所有 lag 均不显著，因子可能无效")
    elif max(sig_lags) <= 5:
        print(f"\n建议：IC 在 lag={max(sig_lags)} 后衰减，建议用 freq='W'（周度换仓）")
    elif max(sig_lags) <= 20:
        print(f"\n建议：IC 在 lag={max(sig_lags)} 后衰减，建议用 freq='ME'（月度换仓）")
    else:
        print(f"\n建议：IC 持续至 lag={max(sig_lags)}，月度换仓足够")

    return result


# ─────────────────────────────────────────────
# 6. 主函数：portfolio sorting（股票池改为半年快照的沪深300+中证500）
# ─────────────────────────────────────────────
def run_portfolio_sort(close_df, adj_df, universe_snapshots,
                       factor_panel_df, trading_dates,
                       start, end, freq='W'):
    """
    universe_snapshots : get_semiannual_snapshots() 的返回结果
                          每个信号日会自动取"该日期之前最近一次"的半年股票池快照
    factor_panel_df     : alpha_func(close_df) 提前算好的日频因子矩阵，主循环直接取截面
    freq                : 换仓频率，建议先跑 calc_ic_decay 后再定
    """
    schedule = build_schedule(trading_dates, start, end, freq)

    all_q_ret = []
    all_factor_panel = []
    all_ret_panel = []

    for signal_date, ret_start, ret_end in schedule:
        # 1. 取该信号日对应的半年股票池快照（800只沪深300+中证500）
        universe_list = get_universe_for_date(signal_date, universe_snapshots)
        if not universe_list:
            continue

        # 2. 因子截面 & 分组
        group = factor_group_from_panel(signal_date, factor_panel_df, universe_list)
        valid_stocks = group.index[group.index.isin(close_df.columns)]
        group = group.loc[valid_stocks]
        if group.empty:
            continue

        # 3. 持仓期收益
        ret_df = calc_adj_returns(close_df, valid_stocks,
                                  ret_start, ret_end)
        if ret_df.empty:
            continue

        # 4. 分组收益
        q_ret = calc_quantile_returns(group, ret_df)
        all_q_ret.append(q_ret)

        # 5. 记录因子面板
        all_factor_panel.append(group['因子值'].reindex(ret_df.columns).rename(signal_date))

        all_ret_panel.append(ret_df)

    combined = pd.concat(all_q_ret).sort_index()
    combined['Q5-Q1'] = combined[5] - combined[1]

    factor_panel_out = pd.DataFrame(all_factor_panel).sort_index()
    ret_panel = pd.concat(all_ret_panel).sort_index()

    return combined, factor_panel_out, ret_panel


FREQ_PERIODS_PER_YEAR = {'D': 252, 'W': 52, 'ME': 12}


def calc_ttest(result_df, col='Q5-Q1', freq='W'):
    periods_per_year = FREQ_PERIODS_PER_YEAR[freq]
    series = result_df[col].dropna()
    t_stat, p_value = stats.ttest_1samp(series, popmean=0)
    return pd.DataFrame({
        '均值': [series.mean()],
        '标准差': [series.std()],
        '年化收益': [series.mean() * periods_per_year],
        '年化夏普': [series.mean() / series.std() * np.sqrt(periods_per_year)],
        't统计量': [t_stat],
        'p值': [p_value],
        '样本数': [len(series)],
    }, index=[col])


def calc_ic(factor_df, ret_df):
    ic_list = {}
    for date in factor_df.index:
        if date not in ret_df.index:
            continue
        f = factor_df.loc[date].dropna()
        r = ret_df.loc[date].dropna()
        common = f.index.intersection(r.index)
        if len(common) < 30:
            continue
        ic_list[date] = f[common].corr(r[common], method='spearman')

    ic = pd.Series(ic_list)
    summary = pd.DataFrame({
        'IC均值': [ic.mean()],
        'IC标准差': [ic.std()],
        'ICIR': [ic.mean() / ic.std()],
        'IC>0占比': [(ic > 0).mean()],
        't统计量': [ic.mean() / (ic.std() / np.sqrt(len(ic)))],
        '样本数': [len(ic)],
    }, index=['因子'])

    return ic, summary


# ─────────────────────────────────────────────
# 7. 行业信息：从 tushare 取（ts_code 本身就是 .SH/.SZ，不需要 SS/SH 转换）
# ─────────────────────────────────────────────
def get_industry_tushare(token):
    """
    从 tushare stock_basic 接口取行业分类
    返回 DataFrame，index=ts_code，列 industry
    """
    pro = ts.pro_api(token)
    df = pro.stock_basic(exchange='', list_status='L', fields='ts_code,name,industry')
    print(df)
    return df.set_index('ts_code')[['industry']]


# ─────────────────────────────────────────────
# 使用示例
# ─────────────────────────────────────────────
if __name__ == "__main__":
    run_mode = 2
    #ALPHA_RANGE = range(30,31)
    ALPHA_RANGE = [75,149]

    #from 量价worldquant101因子库 import ALPHA_REGISTRY
    from 国泰君安191因子库 import ALPHA_REGISTRY
    import pandas as pd

    TUSHARE_TOKEN = 'd12c9df5d1f5406d24ea5e361fe5b93fa0978094c378eb7dea9daff7'  # 替换为你的 token
    DATA_ROOT = r'D:\学习\量化数据\tushare_data'
    SAVE_DIR = r'D:\学习\量化交易\国君191前800回测'
    print("加载全局基础数据...")



    raw_close = load_pivot(f"{DATA_ROOT}\\close_pivot_22-25.csv")
    raw_open = load_pivot(f"{DATA_ROOT}\\open_pivot_22-25.csv")
    raw_high = load_pivot(f"{DATA_ROOT}\\high_pivot_22-25.csv")
    raw_low = load_pivot(f"{DATA_ROOT}\\low_pivot_22-25.csv")
    raw_volume = load_pivot(f"{DATA_ROOT}\\vol_pivot_22-25.csv")
    raw_amount = load_pivot(f"{DATA_ROOT}\\amount_pivot_22-25.csv")
    adj_df = load_pivot(f"{DATA_ROOT}\\adj_factor_pivot_22-25.csv")
    marketcap_df = load_pivot(f"{DATA_ROOT}\\market_value_22-25.csv")
    raw_index = load_pivot(f"{DATA_ROOT}\\000906_SH_22-25.csv")
    raw_pe = load_pivot(f"{DATA_ROOT}\\pe_ttm22-25.csv")

    # 各原始数据源的股票列/交易日不完全一致（如 close 5639 只 vs volume 5651 只），
    # 因子公式里跨字段做 np.where / 布尔赋值时 pandas 不会自动对齐，必须先统一到交集
    _raw_frames = [raw_close, raw_open, raw_high, raw_low, raw_volume, raw_amount, adj_df]
    common_index = _raw_frames[0].index
    common_columns = _raw_frames[0].columns
    for _df in _raw_frames[1:]:
        common_index = common_index.intersection(_df.index)
        common_columns = common_columns.intersection(_df.columns)

    raw_close = raw_close.reindex(index=common_index, columns=common_columns)
    raw_open = raw_open.reindex(index=common_index, columns=common_columns)
    raw_high = raw_high.reindex(index=common_index, columns=common_columns)
    raw_low = raw_low.reindex(index=common_index, columns=common_columns)
    raw_volume = raw_volume.reindex(index=common_index, columns=common_columns)
    raw_amount = raw_amount.reindex(index=common_index, columns=common_columns)
    adj_df = adj_df.reindex(index=common_index, columns=common_columns)
    print(f"已对齐至公共股票池/交易日：{len(common_columns)} 只股票，{len(common_index)} 个交易日")

    # 股票池：半年一次的沪深300+中证500快照
    universe_snapshots = get_semiannual_snapshots(TUSHARE_TOKEN, '2022-01-01', '2025-12-31')

    # 行业信息：直接用 tushare，无需 SS/SH 后缀转换
    industry = get_industry_tushare(TUSHARE_TOKEN)

    print("计算 VWAP 及复权价格...")

    raw_vwap = (raw_amount * 1000) / (raw_volume * 100)

    latest_adj = adj_df.iloc[-1]

    def qfq(raw_df):
        adj_aligned = adj_df.reindex(index=raw_df.index, columns=raw_df.columns)
        return raw_df * adj_aligned / latest_adj

    close_df = qfq(raw_close)
    open_df = qfq(raw_open)
    high_df = qfq(raw_high)
    low_df = qfq(raw_low)
    vwap_df = qfq(raw_vwap)
    volume_df = raw_volume
    amount_df = raw_amount
    index_df = raw_index
    pe_df = raw_pe
    print(f"数据加载完成，close shape: {close_df.shape}")

    for i in ALPHA_RANGE:
        ALPHA_NAME = f"alpha{i:03d}"

        if ALPHA_NAME not in ALPHA_REGISTRY:
            print(f"[{ALPHA_NAME}] 未在注册表中找到，跳过")
            continue

        print(f"\n{'=' * 60}")
        print(f"[{ALPHA_NAME}] 开始计算...")

        dfs = {
            "close": close_df,
            "open": open_df,
            "high": high_df,
            "low": low_df,
            "vwap": vwap_df,
            "volume": volume_df,
            "amount": amount_df,
            "adj_factor": adj_df,
            "industry": industry,
            "cap": marketcap_df,
            "benchmarkindex": index_df,
            "pe": pe_df,
            "universe_snapshots": universe_snapshots,
        }

        try:
            factor_panel_df = ALPHA_REGISTRY[ALPHA_NAME]["func"](dfs)
        except Exception as e:
            print(f"[{ALPHA_NAME}] 因子计算失败: {e}，跳过")
            continue
        print(f"[{ALPHA_NAME}] 完成，shape: {factor_panel_df.shape}")


        if run_mode == 1:
            ic_decay = calc_ic_decay(factor_panel_df, close_df, max_lag=20)

        elif run_mode == 2:
            trading_dates = pd.DatetimeIndex(sorted(close_df.index))
            combined, factor_panel, ret_panel = run_portfolio_sort(
                close_df=close_df,
                adj_df=adj_df,
                universe_snapshots=universe_snapshots,
                factor_panel_df=factor_panel_df,
                trading_dates=trading_dates,
                start='2023-01-01',
                end='2025-12-31',
                freq='W',
            )
            ttest_result = calc_ttest(combined, freq='W')
            ret_panel_aligned = ret_panel.copy()
            ret_panel_aligned.index = factor_panel.index
            ic_series, ic_summary = calc_ic(factor_panel, ret_panel_aligned)
            print(ttest_result)
            print(ic_summary)



            # ── 输出到 Excel ────────────────────────────────────────────
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            import matplotlib

            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            import matplotlib.ticker as mticker
            import io, os

            plt.rcParams['font.sans-serif'] = ['SimHei', 'Microsoft YaHei', 'DejaVu Sans']
            plt.rcParams['axes.unicode_minus'] = False


            os.makedirs(SAVE_DIR, exist_ok=True)
            out_path = os.path.join(SAVE_DIR, f'{ALPHA_NAME}_factor_report.csv')

            wb = openpyxl.Workbook()

            # ── 色彩常量 ────────────────────────────────────────────────
            CLR_HEADER = 'D6E4F0'  # 标题行蓝灰
            CLR_SUBHDR = 'EBF5FB'  # 子标题浅蓝
            CLR_POS = 'E9F7EF'  # 正收益绿底
            CLR_NEG = 'FDEDEC'  # 负收益红底
            CLR_ACCENT = '2980B9'  # 强调色（字体）

            thin = Side(style='thin', color='AAAAAA')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)


            def _hdr_style(cell, bg=CLR_HEADER, bold=True):
                cell.font = Font(name='Arial', bold=bold, size=10, color='1A1A2E')
                cell.fill = PatternFill('solid', fgColor=bg)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border


            def _data_style(cell, fmt=None, bold=False, bg=None):
                cell.font = Font(name='Arial', size=10, bold=bold)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border
                if fmt:
                    cell.number_format = fmt
                if bg:
                    cell.fill = PatternFill('solid', fgColor=bg)


            def _autofit(ws, min_w=10, max_w=20):
                for col in ws.columns:
                    length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
                    ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(length + 2, min_w), max_w)


            # ══════════════════════════════════════════════════════════════
            # Sheet 1: 统计摘要（ttest_result + ic_summary 并排）
            # ══════════════════════════════════════════════════════════════
            ws_stat = wb.active
            ws_stat.title = '统计摘要'
            ws_stat.sheet_view.showGridLines = False


            def _write_df_block(ws, df, start_row, start_col, title):
                """将 DataFrame 写入指定起始单元格，带标题行"""
                # 大标题
                title_cell = ws.cell(row=start_row, column=start_col, value=title)
                title_cell.font = Font(name='Arial', bold=True, size=11, color=CLR_ACCENT)
                title_cell.alignment = Alignment(horizontal='left', vertical='center')
                ws.merge_cells(start_row=start_row, start_column=start_col,
                               end_row=start_row, end_column=start_col + len(df.columns))
                start_row += 1

                # index 列头
                idx_cell = ws.cell(row=start_row, column=start_col, value='')
                _hdr_style(idx_cell)
                for j, col in enumerate(df.columns, start=start_col + 1):
                    c = ws.cell(row=start_row, column=j, value=col)
                    _hdr_style(c)
                start_row += 1

                # 数据行
                for i, (idx, row) in enumerate(df.iterrows()):
                    bg = CLR_SUBHDR if i % 2 == 0 else None
                    idx_c = ws.cell(row=start_row + i, column=start_col, value=str(idx))
                    _data_style(idx_c, bold=True, bg=bg)
                    for j, v in enumerate(row.values, start=start_col + 1):
                        c = ws.cell(row=start_row + i, column=j, value=v)
                        # 数字格式
                        if isinstance(v, float):
                            if abs(v) < 0.1:
                                _data_style(c, fmt='0.0000', bg=bg)
                            else:
                                _data_style(c, fmt='0.0000', bg=bg)
                            # 正负着色
                            if col in ('均值', '年化收益', 'IC均值', 'ICIR') and isinstance(v, float):
                                if v > 0:
                                    c.fill = PatternFill('solid', fgColor=CLR_POS)
                                elif v < 0:
                                    c.fill = PatternFill('solid', fgColor=CLR_NEG)
                        else:
                            _data_style(c, bg=bg)
                return start_row + len(df)


            ws_stat.row_dimensions[1].height = 6  # 顶部留白

            # ttest_result 从第2行第2列开始
            next_row = _write_df_block(ws_stat, ttest_result, start_row=2, start_col=2,
                                       title='▌ 多空组合 t 检验（Q5-Q1）')
            next_row += 1  # 空一行
            # ic_summary 接在下方
            _write_df_block(ws_stat, ic_summary, start_row=next_row, start_col=2,
                            title='▌ IC 汇总统计')

            _autofit(ws_stat)


            # ══════════════════════════════════════════════════════════════
            # 图像辅助函数：生成字节流 → openpyxl Image
            # ══════════════════════════════════════════════════════════════
            def _fig_to_img(fig, dpi=150):
                buf = io.BytesIO()
                fig.savefig(buf, format='png', dpi=dpi, bbox_inches='tight',
                            facecolor=fig.get_facecolor())
                buf.seek(0)
                return openpyxl.drawing.image.Image(buf)


            # ══════════════════════════════════════════════════════════════
            # Sheet 2: IC 时序图（仿石川风格）
            # ══════════════════════════════════════════════════════════════
            ws_ic = wb.create_sheet('IC序列')
            ws_ic.sheet_view.showGridLines = False

            ic_mean = ic_series.mean()
            ic_std = ic_series.std()
            icir = ic_mean / ic_std

            fig, axes = plt.subplots(2, 1, figsize=(14, 8),
                                     gridspec_kw={'height_ratios': [3, 1]},
                                     facecolor='#FAFAFA')
            fig.suptitle(f'{ALPHA_NAME}  IC 序列分析\n'
                         f'IC均值={ic_mean:.4f}  IC_std={ic_std:.4f}  ICIR={icir:.4f}  '
                         f'IC>0占比={(ic_series > 0).mean():.1%}',
                         fontsize=13, fontweight='bold', color='#1A1A2E', y=0.98)

            ax1 = axes[0]
            # 着色区域：正IC绿，负IC红
            ax1.bar(ic_series.index, ic_series.values,
                    color=['#27AE60' if v >= 0 else '#E74C3C' for v in ic_series.values],
                    alpha=0.7, width=5, label='IC')
            # 滚动均值
            ic_roll = ic_series.rolling(12, min_periods=1).mean()
            ax1.plot(ic_roll.index, ic_roll.values, color='#2980B9', lw=1.8,
                     label='12期滚动均值', zorder=5)
            # ±1σ带
            ax1.axhline(ic_mean, color='#8E44AD', lw=1.2, ls='--', label=f'均值={ic_mean:.4f}')
            ax1.axhline(ic_mean + ic_std, color='#AAA', lw=0.8, ls=':', label=f'+1σ={ic_mean + ic_std:.4f}')
            ax1.axhline(ic_mean - ic_std, color='#AAA', lw=0.8, ls=':', label=f'-1σ={ic_mean - ic_std:.4f}')
            ax1.axhline(0, color='#1A1A2E', lw=0.6)
            ax1.set_ylabel('IC (Spearman)', fontsize=10)
            ax1.legend(fontsize=8, loc='upper right', framealpha=0.85)
            ax1.set_facecolor('#F8F9FA')
            ax1.spines['top'].set_visible(False)
            ax1.spines['right'].set_visible(False)
            ax1.grid(axis='y', ls='--', alpha=0.4)

            # 累积IC
            ax2 = axes[1]
            cum_ic = ic_series.cumsum()
            ax2.fill_between(cum_ic.index, cum_ic.values, alpha=0.3, color='#2980B9')
            ax2.plot(cum_ic.index, cum_ic.values, color='#2980B9', lw=1.5)
            ax2.axhline(0, color='#AAA', lw=0.6)
            ax2.set_ylabel('累积IC', fontsize=10)
            ax2.set_facecolor('#F8F9FA')
            ax2.spines['top'].set_visible(False)
            ax2.spines['right'].set_visible(False)
            ax2.grid(axis='y', ls='--', alpha=0.4)

            plt.tight_layout()
            img_ic = _fig_to_img(fig)
            img_ic.anchor = 'B2'
            ws_ic.add_image(img_ic)
            plt.close(fig)

            # ══════════════════════════════════════════════════════════════
            # Sheet 3: 各箱体收益图（仿石川描述性统计图）
            # ══════════════════════════════════════════════════════════════
            ws_port = wb.create_sheet('分组收益')
            ws_port.sheet_view.showGridLines = False

            Q_COLS = [1, 2, 3, 4, 5, 'Q5-Q1']
            Q_LABELS = ['Q1\n（最低）', 'Q2', 'Q3', 'Q4', 'Q5\n（最高）', 'Q5-Q1\n多空']
            PALETTE = ['#E74C3C', '#E67E22', '#95A5A6', '#3498DB', '#27AE60', '#8E44AD']

            fig2, axes2 = plt.subplots(2, 3, figsize=(18, 10), facecolor='#FAFAFA')
            fig2.suptitle(f'{ALPHA_NAME}  分组持仓期收益描述统计',
                          fontsize=14, fontweight='bold', color='#1A1A2E', y=1.01)

            for idx, (col, label, color) in enumerate(zip(Q_COLS, Q_LABELS, PALETTE)):
                ax = axes2[idx // 3][idx % 3]
                ser = combined[col].dropna()

                # ── 背景柱状分布（histogram）
                n, bins, patches = ax.hist(ser, bins=30, color=color, alpha=0.35,
                                           density=False, label='频次')

                # ── 累积分布（右轴）
                ax2r = ax.twinx()
                sorted_vals = np.sort(ser.values)
                cdf = np.arange(1, len(sorted_vals) + 1) / len(sorted_vals)
                ax2r.plot(sorted_vals, cdf, color=color, lw=1.5, alpha=0.8, label='累积分布')
                ax2r.set_ylim(0, 1.15)
                ax2r.yaxis.set_major_formatter(mticker.PercentFormatter(1.0, decimals=0))
                ax2r.set_ylabel('累积占比', fontsize=8, color='#555')
                ax2r.tick_params(axis='y', labelsize=7, colors='#555')
                ax2r.spines['top'].set_visible(False)

                # ── 统计线
                mu = ser.mean()
                med = ser.median()
                q25 = ser.quantile(0.25)
                q75 = ser.quantile(0.75)

                ax.axvline(mu, color='#1A1A2E', lw=1.5, ls='--', label=f'均值 {mu:.2%}')
                ax.axvline(med, color=color, lw=1.2, ls='-', label=f'中位 {med:.2%}')
                ax.axvline(q25, color='#AAA', lw=0.9, ls=':', label=f'Q25 {q25:.2%}')
                ax.axvline(q75, color='#AAA', lw=0.9, ls=':', label=f'Q75 {q75:.2%}')
                ax.axvline(0, color='#888', lw=0.6)

                # 信息框
                info = (f'μ={mu:.2%}  σ={ser.std():.2%}\n'
                        f'Sharpe={mu / ser.std() * np.sqrt(52):.2f}\n'
                        f'n={len(ser)}')
                ax.text(0.03, 0.97, info, transform=ax.transAxes,
                        va='top', ha='left', fontsize=8.5,
                        bbox=dict(boxstyle='round,pad=0.3', fc='white', alpha=0.8, ec=color))

                ax.set_title(label, fontsize=11, fontweight='bold', color=color, pad=6)
                ax.set_xlabel('持仓期收益率', fontsize=9)
                ax.set_ylabel('频次', fontsize=9)
                ax.xaxis.set_major_formatter(mticker.PercentFormatter(1.0, decimals=1))
                ax.set_facecolor('#F8F9FA')
                ax.spines['top'].set_visible(False)
                ax.spines['right'].set_visible(False)
                ax.grid(axis='y', ls='--', alpha=0.3)

                # 合并图例
                lines_left, labels_left = ax.get_legend_handles_labels()
                lines_right, labels_right = ax2r.get_legend_handles_labels()
                ax.legend(lines_left + lines_right, labels_left + labels_right,
                          fontsize=7.5, loc='upper right', framealpha=0.85)

            plt.tight_layout()
            img_port = _fig_to_img(fig2, dpi=130)
            img_port.anchor = 'B2'
            ws_port.add_image(img_port)
            plt.close(fig2)

            # ══════════════════════════════════════════════════════════════
            # Sheet 4: 各箱体累计净值走势图
            # ══════════════════════════════════════════════════════════════
            ws_nav = wb.create_sheet('累计净值')
            ws_nav.sheet_view.showGridLines = False

            fig3, ax3 = plt.subplots(figsize=(14, 6), facecolor='#FAFAFA')
            ax3.set_title(f'{ALPHA_NAME}  各分组累计净值', fontsize=13,
                          fontweight='bold', color='#1A1A2E', pad=10)

            for col, label, color in zip(Q_COLS, Q_LABELS, PALETTE):
                nav = (1 + combined[col].dropna()).cumprod()
                lw = 2.2 if col == 'Q5-Q1' else 1.4
                ls = '--' if col == 'Q5-Q1' else '-'
                ax3.plot(nav.index, nav.values, label=label.replace('\n', ' '),
                         color=color, lw=lw, ls=ls, alpha=0.9)

            ax3.axhline(1, color='#AAA', lw=0.6, ls=':')
            ax3.set_ylabel('净值', fontsize=10)
            ax3.set_facecolor('#F8F9FA')
            ax3.spines['top'].set_visible(False)
            ax3.spines['right'].set_visible(False)
            ax3.grid(axis='y', ls='--', alpha=0.3)
            ax3.legend(fontsize=9, loc='upper left', framealpha=0.85)
            plt.tight_layout()

            img_nav = _fig_to_img(fig3, dpi=150)
            img_nav.anchor = 'B2'
            ws_nav.add_image(img_nav)
            plt.close(fig3)

            # ── 保存 ─────────────────────────────────────────────────────
            wb.save(out_path)
            print(f"\n✓ 报告已保存至: {out_path}")

        else:
            print(f"[{ALPHA_NAME}] run_mode={run_mode} 不合法，请设为 1 或 2")
